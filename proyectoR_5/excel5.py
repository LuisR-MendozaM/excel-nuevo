
# import os
# from datetime import datetime
# from openpyxl import Workbook, load_workbook
# from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# class ExcelUnicoArchivo:
#     def __init__(self, carpeta_datos):
#         # USAR RUTA ABSOLUTA
#         self.carpeta_datos = carpeta_datos
#         self.archivo_principal = os.path.join(carpeta_datos, "Prueba.xlsx")
        
#         print(f"📍 Archivo único: {self.archivo_principal}")
        
#         # Verificar/crear carpeta
#         if not os.path.exists(self.carpeta_datos):
#             print("Carpeta no encontrada")
        
#         # Parámetros (nombre de las hojas)
#         self.parametros = {
#             'temperatura': {'nombre': 'Temperatura', 'unidad': '°C'},
#             'humedad': {'nombre': 'Humedad', 'unidad': '%'},
#             'presion': {'nombre': 'Presión', 'unidad': 'Pa'},
#             'frecuencia': {'nombre': 'Frecuencia', 'unidad': 'Hz'}
#         }
        
#         # SOLO crear el archivo si NO existe, de lo contrario cargarlo
#         self._crear_o_cargar_archivo()
    
#     def _crear_o_cargar_archivo(self):
#         """Crea el archivo solo si no existe, de lo contrario lo carga"""
#         if not os.path.exists(self.archivo_principal):
#             print("Archivo no encontrado")
#         else:
#             print(f"\n📁 CARGANDO ARCHIVO EXISTENTE: {os.path.basename(self.archivo_principal)}")
#             print(f"📍 Ruta: {self.archivo_principal}")
            
#             self._verificar_estructura_archivo()

    
#     def _obtener_hoja_parametro(self, wb, parametro):
#         """Obtiene la hoja correspondiente a un parámetro"""
#         nombre_hoja = self.parametros[parametro]['nombre']
        
#         if nombre_hoja in wb.sheetnames:
#             return wb[nombre_hoja]
#         else:
#             print(f"❌ Hoja '{nombre_hoja}' no encontrada en el archivo")
    
#     def _obtener_posicion_mes(self, mes_num):
#         """Obtiene la posición de la tabla para un mes específico"""
#         # Mapeo de mes a posición
#         mes_a_posicion = {
#             10: (3, 32), # Octubre: A109:B118
#             11: (37, 66), # Noviembre: A121:B130
#             12: (70, 99)  # Diciembre: A133:B142
#         }
#         return mes_a_posicion.get(mes_num, (None, None))
    
#     def _encontrar_fila_vacia(self, ws, inicio, fin):
#         """Encuentra la primera fila vacía en una tabla mensual"""
#         # La primera fila de datos es inicio + 2
#         # (inicio = título, inicio+1 = cabeceras, inicio+2 = primer dato)
#         primera_fila_datos = inicio + 2
        
#         # PRIMERO: Buscar la ÚLTIMA fila con datos en este mes
#         ultima_fila_con_datos = primera_fila_datos - 1  # Empieza en la fila de cabeceras
        
#         for fila in range(primera_fila_datos, fin + 1):
#             if ws[f'A{fila}'].value is not None:
#                 ultima_fila_con_datos = fila
        
#         # Si hay datos, la siguiente fila vacía será después de la última con datos
#         if ultima_fila_con_datos < fin:
#             return ultima_fila_con_datos + 1
#         else:
#             return None  # Tabla llena
    
#     def guardar_dato(self, parametro, valor):
#         """Guarda un dato en la hoja correspondiente del archivo único"""
#         try:
#             if parametro not in self.parametros:
#                 print(f"❌ Parámetro '{parametro}' no válido")
#                 return False
            
#             # Verificar si el archivo existe, si no crearlo
#             if not os.path.exists(self.archivo_principal):
#                 print(f"⚠️ Archivo no encontrado, creando...")
#                 self._crear_archivo_unico()
            
#             # Cargar archivo
#             wb = load_workbook(self.archivo_principal)
            
#             # Obtener hoja del parámetro
#             ws = self._obtener_hoja_parametro(wb, parametro)
            
#             fecha = datetime.now()
#             mes = fecha.month
            
#             # Obtener posición de la tabla del mes
#             inicio, fin = self._obtener_posicion_mes(mes)
#             if inicio is None or fin is None:
#                 print(f"❌ Mes {mes} no tiene posición configurada")
#                 return False
            
#             # Primero verificar si ya existe un registro para HOY
#             fecha_hoy_str = f"{fecha.day:02d}-{fecha.strftime('%b')}-{fecha.year}"
            
#             primera_fila_datos = inicio + 2
#             fila_existente = None
            
#             for fila in range(primera_fila_datos, fin + 1):
#                 if ws[f'A{fila}'].value == fecha_hoy_str:
#                     fila_existente = fila
#                     break
            
#             if fila_existente:
#                 # Actualizar valor existente para hoy
#                 print(f"🔄 Actualizando registro existente para hoy ({fecha_hoy_str})")
#                 fila_vacia = fila_existente
#             else:
#                 # Buscar nueva fila vacía
#                 fila_vacia = self._encontrar_fila_vacia(ws, inicio, fin)
#                 if fila_vacia is None:
#                     print(f"⚠️ Tabla de {self.parametros[parametro]['nombre']} - Mes {mes} está llena")
#                     wb.save(self.archivo_principal)
#                     return False
            
#             # Formatear fecha
#             nombre_mes = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
#                          'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'][mes-1]
#             fecha_str = f"{fecha.day:02d}-{nombre_mes[:3]}-{fecha.year}"
            
#             # Redondear valor
#             try:
#                 valor_redondeado = round(float(valor), 1)
#             except:
#                 valor_redondeado = valor
            
#             # Escribir datos
#             ws[f'A{fila_vacia}'] = fecha_str  # FECHA
#             ws[f'B{fila_vacia}'] = valor_redondeado  # VALOR
            
#             # Guardar cambios
#             wb.save(self.archivo_principal)
            
#             # Mostrar confirmación
#             info = self.parametros[parametro]
#             accion = "Actualizado" if fila_existente else "Guardado"
#             print(f"✅ {accion} {info['nombre']}: {valor_redondeado}{info['unidad']}")
#             print(f"   📅 {fecha_str}")
#             print(f"   📊 Hoja: {info['nombre']} | Tabla: {nombre_mes} | Celda: A{fila_vacia}")
            
#             return True
            
#         except Exception as e:
#             print(f"❌ Error guardando {parametro}: {e}")
#             return False
    
#     def guardar_todos(self, datos):
#         """Guarda todos los parámetros en el archivo único"""
#         print(f"\n💾 GUARDANDO EN ARCHIVO ÚNICO: {os.path.basename(self.archivo_principal)}")
#         print("=" * 60)
        
#         resultados = {}
#         for parametro, valor in datos.items():
#             if parametro in self.parametros:
#                 print(f"\n📋 {self.parametros[parametro]['nombre'].upper()}:")
#                 resultado = self.guardar_dato(parametro, valor)
#                 resultados[parametro] = resultado
        
#         print("\n" + "=" * 60)
#         return resultados
    
#     def ver_estado_archivo(self):
#         """Muestra el estado del archivo único"""
#         try:
#             if not os.path.exists(self.archivo_principal):
#                 return {"error": "Archivo no existe"}
            
#             wb = load_workbook(self.archivo_principal, data_only=True)
            
#             estado = {
#                 'archivo': os.path.basename(self.archivo_principal),
#                 'ruta': self.archivo_principal,
#                 'tamaño_kb': round(os.path.getsize(self.archivo_principal) / 1024, 1),
#                 'hojas': [],
#                 'total_registros': 0
#             }
            
#             # Para cada parámetro/hoja
#             for parametro, info in self.parametros.items():
#                 if info['nombre'] in wb.sheetnames:
#                     ws = wb[info['nombre']]
                    
#                     # Contar registros
#                     registros = 0
#                     for fila in range(1, 143):  # Rango total de todas las tablas
#                         if ws[f'A{fila}'].value is not None and ws[f'A{fila}'].value != "FECHA" and ws[f'A{fila}'].value not in [
#                             'ENERO', 'FEBRERO', 'MARZO', 'ABRIL', 'MAYO', 'JUNIO',
#                             'JULIO', 'AGOSTO', 'SEPTIEMBRE', 'OCTUBRE', 'NOVIEMBRE', 'DICIEMBRE'
#                         ]:
#                             registros += 1
                    
#                     estado['hojas'].append({
#                         'nombre': info['nombre'],
#                         'unidad': info['unidad'],
#                         'registros': registros
#                     })
#                     estado['total_registros'] += registros
            
#             return estado
            
#         except Exception as e:
#             return {"error": str(e)}
    





























# import os
# from datetime import datetime
# from openpyxl import load_workbook

# class ExcelUnicoArchivo:
#     def __init__(self, carpeta_datos):
#         self.archivo = os.path.join(carpeta_datos, "Prueba.xlsm")
#         print(f"📝 Usando: {os.path.basename(self.archivo)}")
        
#         if not os.path.exists(self.archivo):
#             raise FileNotFoundError(f"No se encuentra: {self.archivo}")
    
#     def guardar_dato(self, parametro, valor):
#         """Guarda un dato en la plantilla .xlsm"""
#         try:
#             # Mapeo de parámetros a hojas
#             hojas = {
#                 'temperatura': 'Temperatura',
#                 'humedad': 'Humedad', 
#                 'presion': 'Presion',
#                 'frecuencia': 'Frecuencia'
#             }
            
#             if parametro not in hojas:
#                 return False
            
#             # Cargar con macros
#             wb = load_workbook(self.archivo, keep_vba=True)
#             ws = wb[hojas[parametro]]
            
#             fecha = datetime.now()
#             mes = fecha.month
            
#             # Posiciones por mes (AJUSTA ESTO)
#             posiciones = {
#                 10: (3, 32),   # Octubre
#                 11: (37, 66),  # Noviembre
#                 12: (70, 99)   # Diciembre
#             }
            
#             if mes not in posiciones:
#                 wb.close()
#                 return False
            
#             inicio, fin = posiciones[mes]
            
#             # Buscar fila vacía
#             fila = None
#             for f in range(inicio, fin + 1):
#                 if ws[f'A{f}'].value is None:
#                     fila = f
#                     break
            
#             if fila is None:
#                 wb.close()
#                 return False
            
#             # Escribir fecha (como datetime)
#             ws[f'A{fila}'].value = fecha
#             ws[f'A{fila}'].number_format = 'DD-MMM-YYYY'
            
#             # Escribir valor (tal como viene)
#             ws[f'D{fila}'].value = valor
            
#             # Guardar conservando macros
#             wb.save(self.archivo)
#             wb.close()
            
#             print(f"✓ {hojas[parametro]}: {valor}")
#             return True
            
#         except Exception:
#             return False
    
#     def guardar_todos(self, datos):
#         """Guarda todos los parámetros"""
#         for parametro, valor in datos.items():
#             self.guardar_dato(parametro, valor)
#         print(f"✅ Guardado en {os.path.basename(self.archivo)}")


































import os
from datetime import datetime
from openpyxl import load_workbook

class ExcelUnicoArchivo:
    def __init__(self, carpeta_datos):
        self.archivo = os.path.join(carpeta_datos, "Prueba.xlsm")
        print(f"📝 Usando: {os.path.basename(self.archivo)}")
        
        if not os.path.exists(self.archivo):
            raise FileNotFoundError(f"No se encuentra: {self.archivo}")
        
        # Parámetros para las hojas (AJUSTA los nombres si son diferentes)
        self.parametros = {
            'temperatura': {'nombre': 'Temperatura', 'unidad': '°C'},
            'humedad': {'nombre': 'Humedad', 'unidad': '%'},
            'presion': {'nombre': 'Presion', 'unidad': 'Pa'},
            'frecuencia': {'nombre': 'Frecuencia', 'unidad': 'Hz'}
        }
    
    def guardar_dato(self, parametro, valor):
        """Guarda un dato en la plantilla .xlsm"""
        try:
            # Mapeo de parámetros a hojas
            hojas = {
                'temperatura': 'Temperatura',
                'humedad': 'Humedad', 
                'presion': 'Presion',
                'frecuencia': 'Frecuencia'
            }
            
            if parametro not in hojas:
                print(f"❌ Parámetro no válido: {parametro}")
                return False
            
            # Cargar con macros
            wb = load_workbook(self.archivo, keep_vba=True)
            ws = wb[hojas[parametro]]
            
            # fecha = datetime.now()
            fecha_completa = datetime.now()
            fecha_solo = fecha_completa.date()  # Solo la parte de la fecha
            mes = fecha_completa.month
            
            # Posiciones por mes (AJUSTA ESTO según tu plantilla)
            posiciones = {
                10: (3, 32),   # Octubre
                11: (37, 66),  # Noviembre
                12: (70, 99)   # Diciembre
            }
            
            if mes not in posiciones:
                print(f"❌ Mes {mes} no configurado")
                wb.close()
                return False
            
            inicio, fin = posiciones[mes]
            
            # Buscar fila vacía
            fila = None
            for f in range(inicio, fin + 1):
                if ws[f'A{f}'].value is None:
                    fila = f
                    break
            
            if fila is None:
                print(f"⚠️ Tabla llena para mes {mes}")
                wb.close()
                return False
            
            # Escribir fecha (como datetime)
            ws[f'A{fila}'].value = fecha_solo
            ws[f'A{fila}'].number_format = 'DD-MMM-YYYY'
            
            # Escribir valor en columna D (como número)
            ws[f'D{fila}'].value = valor
            
            # Guardar conservando macros
            wb.save(self.archivo)
            wb.close()
            
            print(f"✓ {hojas[parametro]}: {valor}")
            return True
            
        except Exception as e:
            print(f"❌ Error guardando {parametro}: {e}")
            return False
    
    def guardar_todos(self, datos):
        """Guarda todos los parámetros"""
        print(f"\n💾 Guardando datos...")
        for parametro, valor in datos.items():
            self.guardar_dato(parametro, valor)
        print(f"✅ Guardado en {os.path.basename(self.archivo)}")
    
    # def ver_estado_archivo(self):
    #     """Muestra el estado completo del archivo (BUSCA EN COLUMNA D)"""
    #     try:
    #         if not os.path.exists(self.archivo):
    #             return {"error": "Archivo no existe"}
            
    #         # Cargar solo para lectura
    #         wb = load_workbook(self.archivo, data_only=True)
            
    #         estado = {
    #             'archivo': os.path.basename(self.archivo),
    #             'ruta': self.archivo,
    #             'tamaño_kb': round(os.path.getsize(self.archivo) / 1024, 1),
    #             'total_hojas': len(wb.sheetnames),
    #             'hojas_disponibles': wb.sheetnames,
    #             'hojas': [],
    #             'total_registros': 0
    #         }
            
    #         # Rangos de tablas por mes (DEBE COINCIDIR con guardar_dato)
    #         rangos_meses = {
    #             10: (3, 32),   # Octubre
    #             11: (37, 66),  # Noviembre
    #             12: (70, 99)   # Diciembre
    #         }
            
    #         # Para cada parámetro/hoja
    #         for parametro, info in self.parametros.items():
    #             if info['nombre'] in wb.sheetnames:
    #                 ws = wb[info['nombre']]
    #                 registros_totales = 0
    #                 valores = []
                    
    #                 # Buscar en cada mes configurado
    #                 for mes, (inicio, fin) in rangos_meses.items():
    #                     # Buscar desde inicio hasta fin
    #                     for fila in range(inicio, fin + 1):
    #                         fecha_celda = ws[f'A{fila}'].value
    #                         valor_celda = ws[f'D{fila}'].value  # ← COLUMNA D
                            
    #                         # Si hay ambos, es un registro válido
    #                         if (fecha_celda is not None and 
    #                             valor_celda is not None and
    #                             valor_celda != ""):
                                
    #                             registros_totales += 1
    #                             valores.append(valor_celda)
                    
    #                 # Calcular estadísticas si hay valores
    #                 if valores:
    #                     try:
    #                         # Convertir todos a float para cálculos
    #                         valores_numericos = []
    #                         for v in valores:
    #                             if isinstance(v, (int, float)):
    #                                 valores_numericos.append(float(v))
    #                             elif isinstance(v, str):
    #                                 try:
    #                                     valores_numericos.append(float(v))
    #                                 except:
    #                                     pass
                            
    #                         if valores_numericos:
    #                             ultimo = valores_numericos[-1]
    #                             promedio = sum(valores_numericos) / len(valores_numericos)
    #                             maximo = max(valores_numericos)
    #                             minimo = min(valores_numericos)
    #                         else:
    #                             ultimo = promedio = maximo = minimo = None
    #                     except:
    #                         ultimo = promedio = maximo = minimo = None
    #                 else:
    #                     ultimo = promedio = maximo = minimo = None
                    
    #                 estado['hojas'].append({
    #                     'nombre': info['nombre'],
    #                     'unidad': info['unidad'],
    #                     'registros': registros_totales,
    #                     'ultimo_valor': round(ultimo, 2) if ultimo is not None else None,
    #                     'promedio': round(promedio, 2) if promedio is not None else None,
    #                     'maximo': round(maximo, 2) if maximo is not None else None,
    #                     'minimo': round(minimo, 2) if minimo is not None else None
    #                 })
                    
    #                 estado['total_registros'] += registros_totales
            
    #         wb.close()
    #         return estado
            
    #     except Exception as e:
    #         return {"error": str(e)}
    
    # def imprimir_estado(self):
    #     """Imprime el estado del archivo en formato legible"""
    #     estado = self.ver_estado_archivo()
        
    #     if 'error' in estado:
    #         print(f"❌ Error: {estado['error']}")
    #         return
        
    #     print("\n" + "="*60)
    #     print("📊 ESTADO DEL ARCHIVO EXCEL")
    #     print("="*60)
        
    #     print(f"\n📁 ARCHIVO:")
    #     print(f"   Nombre: {estado['archivo']}")
    #     print(f"   Tamaño: {estado['tamaño_kb']} KB")
    #     print(f"   Total hojas: {estado['total_hojas']}")
    #     print(f"   Hojas disponibles: {', '.join(estado['hojas_disponibles'])}")
        
    #     print(f"\n📈 REGISTROS TOTALES: {estado['total_registros']}")
        
    #     print(f"\n📋 DETALLE POR PARÁMETRO:")
    #     print("-"*60)
        
    #     for hoja in estado['hojas']:
    #         print(f"\n📊 {hoja['nombre'].upper()} ({hoja['unidad']}):")
    #         print(f"   Registros: {hoja['registros']}")
            
    #         if hoja['ultimo_valor'] is not None:
    #             print(f"   Último valor: {hoja['ultimo_valor']}{hoja['unidad']}")
    #             print(f"   Promedio: {hoja['promedio']}{hoja['unidad']}")
    #             print(f"   Máximo: {hoja['maximo']}{hoja['unidad']}")
    #             print(f"   Mínimo: {hoja['minimo']}{hoja['unidad']}")
    #         else:
    #             print(f"   Sin datos registrados")
        
    #     print("\n" + "="*60)
    
    # # Método adicional para ver registros específicos
    # def ver_registros_hoja(self, nombre_hoja, max_registros=10):
    #     """Muestra los últimos registros de una hoja específica"""
    #     try:
    #         wb = load_workbook(self.archivo, data_only=True)
            
    #         if nombre_hoja not in wb.sheetnames:
    #             print(f"❌ Hoja '{nombre_hoja}' no encontrada")
    #             wb.close()
    #             return
            
    #         ws = wb[nombre_hoja]
            
    #         print(f"\n📄 ÚLTIMOS REGISTROS - {nombre_hoja}:")
    #         print("-"*50)
    #         print(f"{'FECHA':<15} {'VALOR':<10} {'CELDA':<10}")
    #         print("-"*50)
            
    #         contador = 0
    #         # Buscar en todas las filas (de abajo hacia arriba)
    #         for fila in range(200, 0, -1):  # Busca de la fila 200 hacia arriba
    #             fecha = ws[f'A{fila}'].value
    #             valor = ws[f'D{fila}'].value  # ← COLUMNA D
                
    #             if fecha is not None and valor is not None and valor != "":
    #                 # Formatear fecha
    #                 if isinstance(fecha, datetime):
    #                     fecha_str = fecha.strftime('%d-%b-%Y')
    #                 else:
    #                     fecha_str = str(fecha)
                    
    #                 print(f"{fecha_str:<15} {valor:<10} D{fila:<9}")
    #                 contador += 1
                    
    #                 if contador >= max_registros:
    #                     break
            
    #         if contador == 0:
    #             print("   No se encontraron registros")
            
    #         wb.close()
            
    #     except Exception as e:
    #         print(f"❌ Error: {e}")